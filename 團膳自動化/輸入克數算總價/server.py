from flask import Flask, request, render_template ## 建立後端伺服器所需要的模組
from openpyxl import Workbook, load_workbook ## 新增到excel所需要的模組 (21行~53行)
from flask import jsonify ##回傳json格式用的
from flask_cors import CORS ##甚麼神魔協定的

app = Flask(__name__)
CORS(app) ## 就是差這個 bug很久

##--------主頁~~-------------
@app.route('/')
def homepage():
    return render_template("home_page.html")
##--------主頁~~-------------





##----------------------(跳轉頁面)來這裡就回傳你新的html ------------------------
@app.route('/add_new_page')
def add_new_page():
    return render_template('add_new_food.html')## 用render_template !! 一定要 !! 把xxx.html 放在templates的資料夾中 !!!! debug超久!!!!!

@app.route('/caculate_total_page')
def caculate_total_page():
    return render_template("caculate_total.html")##render_template 可以執行html (裡面放"檔案路徑 在同一個資料夾就打名稱就好") 不釋放同一個資料夾辣!!! 
##----------------------(跳轉頁面)來這裡就回傳你新的html ------------------------
   





##---------(add new 的表單)按下送出後會來到這裡-----------------
@app.route('/submit')
def submit():


    ##-----------把前端表單的東東存在變數裡--------------
    food_class = request.args.get("class") ## 用 request.args.get 來取得<input name = "name">標籤中 name = "的東東"
    name = request.args.get("name") ## 用 request.args.get 來取得標籤name = "的東東"
    price = request.args.get("price") ## 用 request.args.get 來取得標籤name = "的東東"
    unit = request.args.get("unit") ## 用 request.args.get 來取得標籤name = "的東東"
    ##-----------把前端表單的東東存在變數裡--------------


    ##-----------------------------------------------把拿到的變數新增到excel中------------------------------------------------------------------------------
    wb = load_workbook('add_to_excel_test.xlsx') ##打開在這個資料夾中的excel -- wb=w(ork)b(ook)
    ws = wb[food_class] ## ws=work sheet 就是乳品~豆、魚、蛋、肉類~等等 excel下面的不同工作表(work sheet)
    ##-----------------------------------------------把拿到的變數新增到excel中------------------------------------------------------------------------------


    ##--------------------判斷有幾row---------------------------------   
    row_len = 0                                ##判斷目前row有多長 (資料數目+1 最上面的類別也站一格 )                 ##r1c1 r1c2 r1c3
    for row in range(1,40):                    ##假設不超過40比                                                    ##r2c1 r2c2 r2c3
        if ws['A'+str(row)].value !=None :     ##從A1(食物名稱)到An (不為空的)                                      ##r3c1 r3c2 r3c3
            row_len +=1                                                                                           ##r4c1 r4c2 r4c3 (row max = 4)
        else:                                          
            break
    #print(row_len)
    ##--------------------判斷有幾row--------------------------------- 


    ##-------------檢驗重複 沒有就新增--------------------      
    repeat = False                              
    for row in range (1,row_len+1): ##--------------------從A1開始比對到An
        if(ws['A'+str(row)].value == name):  ##------ws[A1].value 是說 cellA1的值
            repeat = True## ----------------------------A1 A2 A3 ...往下檢驗有沒有重複 
            repeat_row = row
            break                               ##-------一旦有就可以跳出迴圈了
    if(repeat): ##----------------------------如果ws[A3].value (名稱)是 和輸入重複的
            ws['B'+str(repeat_row)].value=price   ##--- 那 B3 得值(價格) 更改成新的
            ws['C'+str(repeat_row)].value=unit  ## ---C3(單位)的值也更新
            repeat = False 
    else:
        ws._current_row = row_len          ##append到現在的格數的下一格
        ws.append([name,price,unit])     ##都沒有重複的話 就在最後新增
    ##-------------檢驗重複 沒有就新增------------------------


    ##----------刪除2~8列---------
    #for i in wb:
    #   i.delete_rows(2,11)
    ##----------刪除2~8列---------


    wb.save('add_to_excel_test.xlsx')    ##存檔

    print('類別 : ' +food_class + ' ' + '名稱 : ' +name  + ' ' + '價格 : '+price   + ' ' + '單位 : ' + unit)
    return render_template('add_new_food.html')  
##--------------------------------------------------把拿到的變數新增到excel中-------------------------------------------------------





##--------------------------------------------------接收前端傳來的類別，回傳此類別的所有名稱------------------------------------------
@app.route('/return_excel_data', methods=['POST'])
def return_excel_data():


    ##-----------把前端傳的東東存在變數裡--------------
    ##food_class = request.args.get("class") ## 用 request.args.get 來取得<input name = "name">標籤中 name = "的東東"
    ##food_class= request.json         ##傳來這個{ data: input_class }
    food_class_JSON = request.get_json()  ##得到{"data" : "類別"}
    food_class = food_class_JSON['data']  ##取出"data"的值 {key : value}
    ##-----------把前端傳的東東存在變數裡--------------


    ##------------打開excel中所需類別的sheet-------------
    wb = load_workbook('add_to_excel_test.xlsx')     ##打開在這個資料夾中的excel -- wb=w(ork)b(ook)
    ws = wb[food_class]                              ## ws=work sheet 就是乳品~豆、魚、蛋、肉類~等等 excel下面的不同工作表(work sheet)
    ##------------打開excel中所需類別的sheet-------------

    row_length_A = 0 ##再取A時 順便計算 A有多少列 等等就B C 九可以和A列數一樣

    ##------------取出所有A列資料-------------
    colA_data = []                             ##list容器放等等找到的A列值                                          ##r1c1 r1c2 r1c3
    for row in range(2,50):                    ##假設不超過50筆                                                    ##r2c1 r2c2 r2c3
        if ws['A'+str(row)].value !=None :     ##從A2(不算開頭的'材料名稱')到An (不為空的)                                      ##r3c1 r3c2 r3c3
            colA_data.append(ws['A'+str(row)].value)        ##如果不為空 那就把這格值加到  colA_data list容器中
            row_length_A  += 1                              ##順便計算 A有多少列                                                            ##r4c1 r4c2 r4c3 (row max = 4)
        else:                                          
            break
    ##print(colA_data) [ '蘋果', '番茄', '芭樂', '玉米'] !! 回傳成功!!!!!!!!!!!!!!!!!!!! 6/23 晚上01.17
    ##------------取出所有A列資料-------------


    ##------------取出所有B列資料-------------
    colB_data = []                             ##list容器放等等找到的B列值                                          ##r1c1 r1c2 r1c3
    for row in range(2,row_length_A+2):                    ## range(1,10)  不包含10    而且A列數沒算到A1 所以+2                  ##r2c1 r2c2 r2c3
        if ws['B'+str(row)].value !=None :     ##從A2(不算開頭的'材料名稱')到An (不為空的)                                      ##r3c1 r3c2 r3c3
            colB_data.append(ws['B'+str(row)].value)        ##如果不為空 那就把這格值加到  colA_data list容器中                                               ##r4c1 r4c2 r4c3 (row max = 4)
        else:                                          
            break
    ##print(colA_data) [ '蘋果', '番茄', '芭樂', '玉米'] !! 回傳成功!!!!!!!!!!!!!!!!!!!! 6/23 晚上01.17
    ##------------取出所有B列資料-------------


    ##------------取出所有C列資料-------------
    colC_data = []                             ##list容器放等等找到的C列值                                          ##r1c1 r1c2 r1c3
    for row in range(2,row_length_A+2):                    ##range(1,10)  不包含10                                                     ##r2c1 r2c2 r2c3
        if ws['C'+str(row)].value !=None :     ##從A2(不算開頭的'材料名稱')到An (不為空的)                                      ##r3c1 r3c2 r3c3
            colC_data.append(ws['C'+str(row)].value)        ##如果不為空 那就把這格值加到  colA_data list容器中                                               ##r4c1 r4c2 r4c3 (row max = 4)
        else:                                          
            break
    ##print(colA_data) [ '蘋果', '番茄', '芭樂', '玉米'] !! 回傳成功!!!!!!!!!!!!!!!!!!!! 6/23 晚上01.17
    ##------------取出所有C列資料-------------


    ##------------回傳資料-------------
    response_data = {
        'result': 'success',
        'colA': colA_data,
        'colB': colB_data,
        'colC': colC_data
    }
    ##------------回傳資料-------------


    return jsonify(response_data) ##回傳json格式
##--------------------------------------------------接收前端傳來的類別，回傳此類別的所有名稱------------------------------------------





##--------單位換算( 給我類別 名稱 需求量 單位 換算後回傳 需求量是多少錢)~~-------------
@app.route('/unit_price_transform')
def unit_price_transform():

    return render_template("home_page.html")
##--------單位換算( 給我類別 名稱 需求量 單位 換算後回傳 需求量是多少錢)~~-------------




if __name__ == '__main__':
    app.run()
    