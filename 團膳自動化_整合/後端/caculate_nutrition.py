from flask import  request
from openpyxl import Workbook, load_workbook ## 新增到excel所需要的模組 (21行~53行)
from openpyxl.styles import Alignment ## 置中用的
from openpyxl.utils import get_column_letter ## 把數字轉字母
from openpyxl.styles import PatternFill ##填充顏色
from openpyxl.formula.translate import Translator ##寫入 翻譯成excel的公式

def caculate_nutrition():

    ##-----------把前端傳的東東存在變數裡--------------
    data_JSON = request.get_json()  ##得到{"data" : "類別"}
    names = data_JSON['names']  ##取出"data"的值 {key : value}
    classes = data_JSON['classes']  ##取出"data"的值 {key : value}
    edible_weights = data_JSON['edible_weights']  ##取出"data"的值 {key : value}
    ##-----------把前端傳的東東存在變數裡--------------

    ##打開計算營養成分 excel
    wb = load_workbook('計算營養成分模板.xlsx') ##打開在這個資料夾中的excel -- wb=w(ork)b(ook)
    ws = wb["菜單設計及營養成份加總"] ## ws=work sheet 

    # 將類別填充到 A6 至 結束
    for index, item in enumerate(classes, start=6):
        ws[f'A{index}'] = item

    # 將名稱填充到 B6 至 結束
    for index, item in enumerate(names, start=6):
        ws[f'B{index}'] = item

    # 將可食重量到 C6 至 結束
    for index, item in enumerate(edible_weights, start=6):
        ws[f'C{index}'] = item

    #儲存工作簿
    wb.save('計算營養成分結果.xlsx')

    