from flask import  request
from flask import jsonify ##回傳json格式用的
from openpyxl import Workbook, load_workbook ## 新增到excel所需要的模組 (21行~53行)

def get_ingrident_name_data():


    ##-----------把前端傳的東東存在變數裡--------------
    food_class_JSON = request.get_json()  ##得到{"data" : "類別"}
    food_class = food_class_JSON['data']  ##取出"data"的值 {key : value}
    ##-----------把前端傳的東東存在變數裡--------------


    ##------------打開excel中所需類別的sheet-------------
    wb = load_workbook('食材庫.xlsx')     ##打開在這個資料夾中的excel -- wb=w(ork)b(ook) 
    ws = wb[food_class]                              ## ws=work sheet 就是乳品~豆、魚、蛋、肉類~等等 excel下面的不同工作表(work sheet)
    ## 好奇怪 原來excel黨要放在更外面的資料夾??
    ##------------打開excel中所需類別的sheet-------------

    row_length_A = 0 ##再取A時 順便計算 A有多少列 等等就B C 九可以和A列數一樣

    ##------------取出所有A列資料-------------
    colA_data = []                             ##list容器放等等找到的A列值                                          ##r1c1 r1c2 r1c3
    for row in range(2,500):                    ##假設不超過500筆                                                    ##r2c1 r2c2 r2c3
        if ws['A'+str(row)].value !=None :     ##從A2(不算開頭的'材料名稱')到An (不為空的)                                      ##r3c1 r3c2 r3c3
            colA_data.append(ws['A'+str(row)].value)        ##如果不為空 那就把這格值加到  colA_data list容器中
            row_length_A  += 1                              ##順便計算 A有多少列                                                            ##r4c1 r4c2 r4c3 (row max = 4)
        else:                                          
            break
    ##print(colA_data) [ '蘋果', '番茄', '芭樂', '玉米'] !! 回傳成功!!!!!!!!!!!!!!!!!!!! 6/23 晚上01.17
    ##------------取出所有A列資料-------------


    ##------------回傳資料-------------
    response_data = {
        'result': 'success',
        'colA': colA_data
    }
    ##------------回傳資料-------------

    return jsonify(response_data) ##回傳json格式