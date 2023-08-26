from flask import  request
from openpyxl import Workbook, load_workbook ## 新增到excel所需要的模組 (21行~53行)
from openpyxl.styles import Alignment ## 置中用的
from openpyxl.utils import get_column_letter ## 把數字轉字母
from openpyxl.styles import PatternFill ##填充顏色
from openpyxl.formula.translate import Translator ##寫入 翻譯成excel的公式
import xlwings as xw # 強制執行公式計算用的


all_nutrition_datas = {}

def caculate_all_nutrition():

    ##-----------把前端傳的東東存在變數裡--------------
    data_JSON = request.get_json()  ##得到{"data" : "類別"}
    names = data_JSON['names']  ##取出"data"的值 {key : value}
    classes = data_JSON['classes']  ##取出"data"的值 {key : value}
    edible_weights = data_JSON['edible_weights']  ##取出"data"的值 {key : value}
    meal_time = data_JSON['which_mealtime']
    ##-----------把前端傳的東東存在變數裡--------------

    ##打開計算營養成分 excel
    wb = load_workbook('計算營養成分模板.xlsx') #data_only=True參數，這樣當你從包含公式的單元格中讀取時，將返回計算結果而不是公式本身。


    # 看要打開哪一個sheet
    if (meal_time=='AETMT_1_table'):
        ws = wb["早餐"] ## ws=work sheet 
    elif (meal_time=='AETMT_2_table'):
        ws = wb["早點"] ## ws=work sheet 
    elif (meal_time=='AETMT_3_table'):
        ws = wb["午餐"] ## ws=work sheet 
    elif (meal_time=='AETMT_4_table'):
        ws = wb["午點"] ## ws=work sheet 
    elif (meal_time=='AETMT_5_table'):
        ws = wb["晚餐"] ## ws=work sheet 
    elif (meal_time=='AETMT_6_table'):
        ws = wb["晚點"] ## ws=work sheet 



    # 將類別填充到 A6 至 結束
    for index, item in enumerate(classes, start=6):
        ws[f'A{index}'] = item

    # 將名稱填充到 B6 至 結束
    for index, item in enumerate(names, start=6):
        ws[f'B{index}'] = item

    # 將可食重量到 C6 至 結束
    for index, item in enumerate(edible_weights, start=6):
        # 不是空字串
        if(item) : 
            ws[f'C{index}'] = float(item)
        else:
            print("ERROR : 可食重量為空字串")
            break

    #儲存工作簿
    wb.save('計算營養成分結果.xlsx')

    # ----- 強制計算 (把公式算出答案)--------
    # 打開Excel工作簿
    wb = xw.Book('計算營養成分結果.xlsx')
    # 強制計算所有公式
    wb.app.calculate()
    # 保存並關閉工作簿
    wb.save('計算營養成分結果.xlsx')
    wb.close()
    # ----- 強制計算--------

    wb = load_workbook('計算營養成分結果.xlsx',data_only=True) #data_only=True參數，這樣當你從包含公式的單元格中讀取時，將返回計算結果而不是公式本身。

    # 看要打開哪一個sheet
    if (meal_time=='AETMT_1_table'):
        ws = wb["早餐"] ## ws=work sheet 
    elif (meal_time=='AETMT_2_table'):
        ws = wb["早點"] ## ws=work sheet 
    elif (meal_time=='AETMT_3_table'):
        ws = wb["午餐"] ## ws=work sheet 
    elif (meal_time=='AETMT_4_table'):
        ws = wb["午點"] ## ws=work sheet 
    elif (meal_time=='AETMT_5_table'):
        ws = wb["晚餐"] ## ws=work sheet 
    elif (meal_time=='AETMT_6_table'):
        ws = wb["晚點"] ## ws=work sheet 

    # Step 2: 從E1到DD1獲取營養名稱值   ["熱量 (kcal)"	"粗蛋白 (g)"	"粗脂肪 (g)"	"總碳水化合物 (g)"	"糖質總量 (g)"	"水分 (g)"	"膽固醇 (mg)"	"膳食纖維 (g)"]
    nutrition_names = [ws.cell(row=1, column=i).value for i in range(5, 109)]  # 5是E的列號，110是DD的列號加1
    # Step 2: 從E3到DD3獲取值  [7.2 	0.0 	0.0 	0.0 	0.0 	0.0 	0.0 	0.0 ]
    nutrition_values = [round(ws.cell(row=3, column=i).value, 2) for i in range(5, 109)]  # 5是E的列號，110是DD的列號加1 取道小數點第二位

    # 把所有資料按照 早餐 午餐...分配 {早餐 : [all_nutrition_names , all_nutrition_datas], 午餐 : [all_nutrition_names , all_nutrition_datas], 晚餐 : [all_nutrition_names , all_nutrition_datas]}
    meal_time_name = ws.title
    all_nutrition_datas[meal_time_name] = [nutrition_names , nutrition_values] 

    return all_nutrition_datas
