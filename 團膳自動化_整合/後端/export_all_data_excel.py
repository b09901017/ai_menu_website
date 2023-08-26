from flask import  request
from openpyxl import Workbook, load_workbook ## 新增到excel所需要的模組 (21行~53行)
from openpyxl.styles import Alignment ## 置中用的
from openpyxl.utils import get_column_letter ## 把數字轉字母
from openpyxl.styles import PatternFill ##填充顏色
from openpyxl.formula.translate import Translator ##寫入 翻譯成excel的公式
from itertools import chain  # 把二微陣列展平 [[1,2,3],[1,2,3],[1,2,3],[4,5,6]] = [1,2,3,1,2,3,1,2,3,4,5,6] ! 8/22 cool~~~


# // 回傳的東東 舉例
#     // {
#     //     ratio_table:(4) ['1400', '20%', '30%', '50%']
#     //     EX_array: (10) ['0', '1', '5', '9', '0', '0', '1', '4', 2'', '0']
#     //     AETMT_meal_time: (6) [Array(10), Array(10), Array(10), Array(10), Array(10), Array(10)]
                                # 0: (10) ['', '', '', '', '', '', '', '', '', '']
                                # 1: (10) ['', '', '', '', '', '', '', '', '', '']
                                # 2: (10) ['', '', '', '', '', '', '', '', '', '']
                                # 3: (10) ['', '', '', '', '', '', '', '', '', '']
                                # 4: (10) ['', '', '', '', '', '', '', '', '', '']
                                # 5: (10) ['', '', '', '', '', '', '', '', '', '']

#     //     dish_data: {meal_1: {…}, meal_2: {…}, meal_3: {…}, meal_4: {…}, meal_5: {…}, …}
                                # meal_1 : {dish_1: {…}, dish_2: {…}}
                                            # dish_1 :
                                                    # class_1 : (3) ['穀物類', '穀物類', '穀物類']
                                                    # class_2 : (3) ['奶 全脂', '奶 全脂', '奶 全脂']
                                                    # edible_weight : (3) ['', '', '']
                                                    # ex : (3) ['', '', '']
                                                    # ingredient :(3) ['', '', '']
                                                    # name ： ＂＂
                                # meal_2 : {dish_1: {…}, dish_2: {…}}
                                # meal_3 : {dish_1: {…}, dish_2: {…}}
                                # meal_4 : {dish_1: {…}, dish_2: {…}}
                                # meal_5 : {dish_1: {…}, dish_2: {…}}
                                # meal_6 : {dish_1: {…}, dish_2: {…}}
#           }
# 多傳入 all_nutrition_datas 8/24 放入營養成分
def export_all_data_excel(all_nutrition_datas) :

    ##-----------把前端傳來的東東存在變數裡--------------
    data_json = request.get_json()

    ratio_table = data_json['ratio_table']
    EX_array = data_json['EX_array']
    AETMT_meal_time = data_json['AETMT_meal_time']
    dish_data = data_json['dish_data']

    total_kcal = ratio_table[0]
    sugar_ratio = ratio_table[1]
    lipid_ratio = ratio_table[2]
    protain_ratio = ratio_table[3]
    ##-----------把前端傳來的東東存在變數裡--------------

    ##------------打開excel中所需類別的sheet-------------
    wb = load_workbook('自動化表格試做.xlsx')     ##打開在這個資料夾中的excel -- wb=w(ork)b(ook) 


    ##-----------三大營養數比例-----------------------
    ws = wb['三大營養素比例']
    ws['B1'] = total_kcal + "大卡"
    ws['B3'] = sugar_ratio
    ws['B4'] = lipid_ratio
    ws['B5'] = protain_ratio
    ##-----------三大營養數比例-----------------------


    ##-----------份數分配-----------------------
    ws = wb['Meal pattern']
    # 放入C3~C12
    for index, item in enumerate(EX_array):
        # 如果是空字串 就不能用int() 強制轉換 幹~ 原來是境製糖 還沒放數字個關西
        if item.isdigit():
            ws.cell(row=3 + index, column=3, value=float(item))
        else:
            # 如果是空白那就是0
            ws.cell(row=3 + index, column=3, value=0)
    ##-----------份數分配-----------------------


    ##-----------餐次分配-----------------------
    ws = wb['Meal pattern']
    # 放入C3~C12
    # 對於數據中的每個子陣列
    for i, meal_time in enumerate(AETMT_meal_time):
        for j, item in enumerate(meal_time):
            if item.isdigit():
                # 將每個元素放入正確的儲存格位置 H是8 H~M
                ws.cell(row=3 + j, column=8 + i, value=float(item))
            else:
                # 如果是空白""那就 放入 - 
                ws.cell(row=3 + j, column=8 + i, value="-")
    ##-----------餐次分配-----------------------


    ## ------------- 菜單 + 成本 + 營養成分 -------------------------------------
    # region 
     
    # region 早餐
    # -----------早餐菜單 ---------------------------
    # region 
    ws = wb["早餐  菜單+成本"]
    meal_1 = dish_data['meal_1']

    # 使用列表解析找到找餐中所有以 "dish_" 開頭的鍵
    dish_keys = [key for key in meal_1.keys() if key.startswith("dish_")]

    # 計算找到的菜名的數量
    number_of_dishes = len(dish_keys)

    # 使用列表解析取得每一個dish_n的name class 等等屬性 放到陣列中
    dish_names = [dish["name"] for dish in meal_1.values()] #["name1",name2,name3...]
    dish_ingredients = [dish["ingredient"] for dish in meal_1.values()] # [[],[],[]...]
    dish_class_1s = [dish["class_1"] for dish in meal_1.values()]
    dish_class_2s = [dish["class_2"] for dish in meal_1.values()]
    dish_exs = [dish["ex"] for dish in meal_1.values()]
    dish_edible_weights = [dish["edible_weight"] for dish in meal_1.values()]
    # 8/27 如果沒有任何一道菜 (反正部會進到for迴圈)  ws.merge_cells(start_row=5, start_column=1, end_row=start_row-1, end_column=1) 使end_row > start_row=5
    if(len(dish_names) == 0):
        start_row =6 #為了等等被減1
    else:
        # 從B5開始一直往下田dish
        start_row = 5
    
    for i in range(number_of_dishes):
        # 開始一個一個放 (仙取的這個dish有多少材料)
        ingredients_length = len(dish_ingredients[i])
        # 從B5~ 材料數+5-1 (但如果材料數是0 (沒有任何菜名) ) 那就不要減一(因為last = start+0-1 會比start小) ERROR : 4 must be greater than 5
        if(ingredients_length==0):
            last_row = start_row 
        else:
            last_row = start_row + ingredients_length -1
        # 合併菜名 並 放入菜名 (先unmerge)
        ws.cell(row=start_row, column=2, value=dish_names[i])
        ws.merge_cells(start_row=start_row, start_column=2, end_row=last_row, end_column=2)
        # 放材料名稱 C = 3
        for index, value in enumerate(dish_ingredients[i]):
            ws.cell(row=start_row + index, column=3, value=value)
        # 放入類別1 D = 4
        for index, value in enumerate(dish_class_1s[i]):
            ws.cell(row=start_row + index, column=4, value=value)
        # 放類別2 E = 5
        for index, value in enumerate(dish_class_2s[i]):
            ws.cell(row=start_row + index, column=5, value=value)
        # 放EX F = 6
        for index, value in enumerate(dish_exs[i]):
            ws.cell(row=start_row + index, column=6, value=value)
        # 放可食重量 G = 7 (要是數字)
        for index, value in enumerate(dish_edible_weights[i]):
            if value: # 判斷不是空字串 有值就是TRUE
                value = float(value)
                ws.cell(row=start_row + index, column=7, value=value)
            else : 
                ws.cell(row=start_row + index, column=7, value=value)
        # 把作法合併
        ws.merge_cells(start_row=start_row, start_column=8, end_row=last_row, end_column=8)
        # 最後把起始start_row 加上填的材料數
        start_row += ingredients_length

    # 在迴圈外 把餐別加上去 
    ws['A5'] = "早餐"
    ws.merge_cells(start_row=5, start_column=1, end_row=start_row-1, end_column=1)
   # endregion
    # -----------早餐菜單 ---------------------------

    # ----------- 早餐成本 (把食材篩選出來而已 一樣的就不重打)--------------------------
    # region 

    # 使用 itertools.chain 來展平列表 dish_ingredients 已經是 meal_1取出來的二維陣列[[],[],[]]
    dish_ingredients = list(chain(*dish_ingredients))

    # 取出不一樣的
    unique_ingredients = []
    for item in dish_ingredients:
        if item not in unique_ingredients:
            unique_ingredients.append(item)

    # 放入excel
    for index, value in enumerate(unique_ingredients):
            ws.cell(row=6 + index, column=10, value=value)

    # endregion
    # ----------- 早餐成本----------------------------------------------------------

    # ----------- 早餐營養成分 --------------------------
    # region 
    # 遍歷 S4~Z4  營養素分析
    for i in range(19, 27):

        # 如果all_nutrition_datas = None 如同server.py 初始化的 (早餐還沒按下計算營養按鈕)
        if all_nutrition_datas is None:
            print("早餐還沒按下計算營養按鈕 (key = '早餐' 沒發現)")
            break
        # 阿如果all_nutrition_datas["早餐"] 有東西 那就執行下面
        else:
            # 去除 S4_value 中的所有空格和換行符：
            cleaned_value = ws.cell(row=4, column=i).value.replace("\n", "").replace(" ", "")
            # {早餐 : [all_nutrition_names , all_nutrition_values], 午餐 : [ [] , [] ]}
            try:
                # 找到S4"熱量 (kcal)" 所匹配到的 營養成分結果的index
                compare_name_index = all_nutrition_datas["早餐"][0].index(cleaned_value)
                # 把早餐的熱量 (or 其他的營養成分) 有多少抓過來
                this_nutrition_value = all_nutrition_datas["早餐"][1][compare_name_index]
                # 放入S5
                ws.cell(row=5, column=i,value = this_nutrition_value)
            except ValueError:
                ws.cell(row=5, column=i,value = "not found")
            except KeyError:
                print("KEYERROR : 早餐~~")
                break
    # 遍歷 S12~AA12 礦物質分析
    for i in range(19, 28):

        # 如果all_nutrition_datas = None 如同server.py 初始化的 (早餐還沒按下計算營養按鈕)
        if all_nutrition_datas is None:
            print("早餐還沒按下計算營養按鈕 (key = '早餐' 沒發現)")
            break
        # 阿如果all_nutrition_datas["早餐"] 有東西 那就執行下面
        else:
            # 去除 S4_value 中的所有空格和換行符：
            cleaned_value = ws.cell(row=12, column=i).value.replace("\n", "").replace(" ", "")
            # {早餐 : [all_nutrition_names , all_nutrition_values], 午餐 : [ [] , [] ]}
            try:
                # 找到S4"熱量 (kcal)" 所匹配到的 營養成分結果的index
                compare_name_index = all_nutrition_datas["早餐"][0].index(cleaned_value)
                # 把早餐的熱量 (or 其他的營養成分) 有多少抓過來
                this_nutrition_value = all_nutrition_datas["早餐"][1][compare_name_index]
                # 放入S5
                ws.cell(row=13, column=i,value = this_nutrition_value)
            except ValueError:
                ws.cell(row=13, column=i,value = "not found")
            except KeyError:
                print("KEYERROR : 早餐~~")
                break
    # 遍歷 S20~AA20 維生素分析
    for i in range(19, 28):

        # 如果all_nutrition_datas = None 如同server.py 初始化的 (早餐還沒按下計算營養按鈕)
        if all_nutrition_datas is None:
            print("早餐還沒按下計算營養按鈕 (key = '早餐' 沒發現)")
            break
        # 阿如果all_nutrition_datas["早餐"] 有東西 那就執行下面
        else:
            # 去除 S4_value 中的所有空格和換行符：
            cleaned_value = ws.cell(row=20, column=i).value.replace("\n", "").replace(" ", "")
            # {早餐 : [all_nutrition_names , all_nutrition_values], 午餐 : [ [] , [] ]}
            try:
                # 找到S4"熱量 (kcal)" 所匹配到的 營養成分結果的index
                compare_name_index = all_nutrition_datas["早餐"][0].index(cleaned_value)
                # 把早餐的熱量 (or 其他的營養成分) 有多少抓過來
                this_nutrition_value = all_nutrition_datas["早餐"][1][compare_name_index]
                # 放入S5
                ws.cell(row=21, column=i,value = this_nutrition_value)
            except ValueError:
                ws.cell(row=21, column=i,value = "not found")
            except KeyError:
                print("KEYERROR : 早餐~~")
                break

    # endregion
    # ----------- 早餐營養成分----------------------------------------------------------
    # endregion

    # region 早點
    # -----------早點菜單 ---------------------------
    # region 
    ws = wb["早點  菜單+成本"]
    meal_1 = dish_data['meal_2']

    # 使用列表解析找到找餐中所有以 "dish_" 開頭的鍵
    dish_keys = [key for key in meal_1.keys() if key.startswith("dish_")]

    # 計算找到的菜名的數量
    number_of_dishes = len(dish_keys)

    # 使用列表解析取得每一個dish_n的name class 等等屬性 放到陣列中
    dish_names = [dish["name"] for dish in meal_1.values()] #["name1",name2,name3...]
    dish_ingredients = [dish["ingredient"] for dish in meal_1.values()] # [[],[],[]...]
    dish_class_1s = [dish["class_1"] for dish in meal_1.values()]
    dish_class_2s = [dish["class_2"] for dish in meal_1.values()]
    dish_exs = [dish["ex"] for dish in meal_1.values()]
    dish_edible_weights = [dish["edible_weight"] for dish in meal_1.values()]
    # 8/27 如果沒有任何一道菜 (反正部會進到for迴圈)  ws.merge_cells(start_row=5, start_column=1, end_row=start_row-1, end_column=1) 使end_row > start_row=5
    if(len(dish_names) == 0):
        start_row =6 #為了等等被減1
    else:
        # 從B5開始一直往下田dish
        start_row = 5

    # 如果沒有任何一道菜 8/27 那迴圈連進去都不會進去 
    for i in range(number_of_dishes):
        # 開始一個一個放 (仙取的這個dish有多少材料)
        ingredients_length = len(dish_ingredients[i])
        # 從B5~ 材料數+5-1 (但如果材料數是0 (沒有任何菜名) ) 那就不要減一(因為last = start+0-1 會比start小) ERROR : 4 must be greater than 5
        if(ingredients_length==0):
            last_row = start_row 
        else:
            last_row = start_row + ingredients_length -1
        # 合併菜名 並 放入菜名
        ws.cell(row=start_row, column=2, value=dish_names[i])
        ws.merge_cells(start_row=start_row, start_column=2, end_row=last_row, end_column=2)
        # 放材料名稱 C = 3
        for index, value in enumerate(dish_ingredients[i]):
            ws.cell(row=start_row + index, column=3, value=value)
        # 放入類別1 D = 4
        for index, value in enumerate(dish_class_1s[i]):
            ws.cell(row=start_row + index, column=4, value=value)
        # 放類別2 E = 5
        for index, value in enumerate(dish_class_2s[i]):
            ws.cell(row=start_row + index, column=5, value=value)
        # 放EX F = 6
        for index, value in enumerate(dish_exs[i]):
            ws.cell(row=start_row + index, column=6, value=value)
        # 放可食重量 G = 7
        for index, value in enumerate(dish_edible_weights[i]):
            if (value): # 判斷不是空字串
                print(start_row + index)
                value = float(value)
                ws.cell(row=start_row + index, column=7, value=value)
            else : 
                ws.cell(row=start_row + index, column=7, value=value)

        # 把作法合併
        ws.merge_cells(start_row=start_row, start_column=8, end_row=last_row, end_column=8)
        # 最後把起始start_row 加上填的材料數  ( 8/27 如果一道菜都沒有 那start_row就等於start_row = 5 沒差 如果沒有任何一道菜 那根本連for迴圈都不會進來)
        start_row += ingredients_length

    # 在迴圈外 把餐別加上去 
    ws['A5'] = "早點"
    ws.merge_cells(start_row=5, start_column=1, end_row=start_row-1, end_column=1)
   # endregion
    # -----------早點菜單 ---------------------------

    # ----------- 早點成本 (把食材篩選出來而已 一樣的就不重打)--------------------------
    # region 

    # 使用 itertools.chain 來展平列表 dish_ingredients 已經是 meal_1取出來的二維陣列[[],[],[]]
    dish_ingredients = list(chain(*dish_ingredients))

    # 取出不一樣的
    unique_ingredients = []
    for item in dish_ingredients:
        if item not in unique_ingredients:
            unique_ingredients.append(item)

    # 放入excel
    for index, value in enumerate(unique_ingredients):
            ws.cell(row=6 + index, column=10, value=value)
            
    # endregion
    # ----------- 早點成本----------------------------------------------------------

    # ----------- 早點營養成分 --------------------------
    # region 
    # 遍歷 S4~Z4  營養素分析
    for i in range(19, 27):

        # 如果all_nutrition_datas = None 如同server.py 初始化的 (早餐還沒按下計算營養按鈕)
        if all_nutrition_datas is None:
            print("早點還沒按下計算營養按鈕 (key = '早點' 沒發現)")
            break
        # 阿如果all_nutrition_datas["早點"] 有東西 那就執行下面
        else:
            # 去除 S4_value 中的所有空格和換行符：
            cleaned_value = ws.cell(row=4, column=i).value.replace("\n", "").replace(" ", "")
            # {早餐 : [all_nutrition_names , all_nutrition_values], 午餐 : [ [] , [] ]}
            try:
                # 找到S4"熱量 (kcal)" 所匹配到的 營養成分結果的index
                compare_name_index = all_nutrition_datas["早點"][0].index(cleaned_value)
                # 把早餐的熱量 (or 其他的營養成分) 有多少抓過來
                this_nutrition_value = all_nutrition_datas["早點"][1][compare_name_index]
                # 放入S5
                ws.cell(row=5, column=i,value = this_nutrition_value)
            except ValueError:
                ws.cell(row=5, column=i,value = "not found")
            except KeyError:
                print("KEYERROR : 早點~~")
                break
            
    # 遍歷 S12~AA12 礦物質分析
    for i in range(19, 28):

        # 如果all_nutrition_datas = None 如同server.py 初始化的 (早餐還沒按下計算營養按鈕)
        if all_nutrition_datas is None:
            print("早點還沒按下計算營養按鈕 (key = '早點' 沒發現)")
            break
        # 阿如果all_nutrition_datas["早餐"] 有東西 那就執行下面
        else:
            # 去除 S4_value 中的所有空格和換行符：
            cleaned_value = ws.cell(row=12, column=i).value.replace("\n", "").replace(" ", "")
            # {早餐 : [all_nutrition_names , all_nutrition_values], 午餐 : [ [] , [] ]}
            try:
                # 找到S4"熱量 (kcal)" 所匹配到的 營養成分結果的index
                compare_name_index = all_nutrition_datas["早點"][0].index(cleaned_value)
                # 把早餐的熱量 (or 其他的營養成分) 有多少抓過來
                this_nutrition_value = all_nutrition_datas["早點"][1][compare_name_index]
                # 放入S5
                ws.cell(row=13, column=i,value = this_nutrition_value)
            except ValueError:
                ws.cell(row=13, column=i,value = "not found")
            except KeyError:
                print("KEYERROR : 早點~~")
                break
    # 遍歷 S20~AA20 維生素分析
    for i in range(19, 28):

        # 如果all_nutrition_datas = None 如同server.py 初始化的 (早餐還沒按下計算營養按鈕)
        if all_nutrition_datas is None:
            print("早餐還沒按下計算營養按鈕 (key = '早餐' 沒發現)")
            break
        # 阿如果all_nutrition_datas["早餐"] 有東西 那就執行下面
        else:
            # 去除 S4_value 中的所有空格和換行符：
            cleaned_value = ws.cell(row=20, column=i).value.replace("\n", "").replace(" ", "")
            # {早餐 : [all_nutrition_names , all_nutrition_values], 午餐 : [ [] , [] ]}
            try:
                # 找到S4"熱量 (kcal)" 所匹配到的 營養成分結果的index
                compare_name_index = all_nutrition_datas["早點"][0].index(cleaned_value)
                # 把早餐的熱量 (or 其他的營養成分) 有多少抓過來
                this_nutrition_value = all_nutrition_datas["早點"][1][compare_name_index]
                # 放入S5
                ws.cell(row=21, column=i,value = this_nutrition_value)
            except ValueError:
                ws.cell(row=21, column=i,value = "not found")
            except KeyError:
                print("KEYERROR : 早點~~")
                break

    # endregion
    # ----------- 早點營養成分----------------------------------------------------------
# endregion

    # region 午餐
    # -----------午餐菜單 ---------------------------
    # region 
    ws = wb["午餐  菜單+成本"]
    meal_1 = dish_data['meal_3']

    # 使用列表解析找到找餐中所有以 "dish_" 開頭的鍵
    dish_keys = [key for key in meal_1.keys() if key.startswith("dish_")]

    # 計算找到的菜名的數量
    number_of_dishes = len(dish_keys)

    # 使用列表解析取得每一個dish_n的name class 等等屬性 放到陣列中
    dish_names = [dish["name"] for dish in meal_1.values()] #["name1",name2,name3...]
    dish_ingredients = [dish["ingredient"] for dish in meal_1.values()] # [[],[],[]...]
    dish_class_1s = [dish["class_1"] for dish in meal_1.values()]
    dish_class_2s = [dish["class_2"] for dish in meal_1.values()]
    dish_exs = [dish["ex"] for dish in meal_1.values()]
    dish_edible_weights = [dish["edible_weight"] for dish in meal_1.values()]
    # 8/27 如果沒有任何一道菜 (反正部會進到for迴圈)  ws.merge_cells(start_row=5, start_column=1, end_row=start_row-1, end_column=1) 使end_row > start_row=5
    if(len(dish_names) == 0):
        start_row =6 #為了等等被減1
    else:
        # 從B5開始一直往下田dish
        start_row = 5

    for i in range(number_of_dishes):
        # 開始一個一個放 (仙取的這個dish有多少材料)
        ingredients_length = len(dish_ingredients[i])
        # 從B5~ 材料數+5-1
        last_row = start_row + ingredients_length -1
        # 合併菜名 並 放入菜名
        ws.cell(row=start_row, column=2, value=dish_names[i])
        ws.merge_cells(start_row=start_row, start_column=2, end_row=last_row, end_column=2)
        # 放材料名稱 C = 3
        for index, value in enumerate(dish_ingredients[i]):
            ws.cell(row=start_row + index, column=3, value=value)
        # 放入類別1 D = 4
        for index, value in enumerate(dish_class_1s[i]):
            ws.cell(row=start_row + index, column=4, value=value)
        # 放類別2 E = 5
        for index, value in enumerate(dish_class_2s[i]):
            ws.cell(row=start_row + index, column=5, value=value)
        # 放EX F = 6
        for index, value in enumerate(dish_exs[i]):
            ws.cell(row=start_row + index, column=6, value=value)
        # 放可食重量 G = 7
        for index, value in enumerate(dish_edible_weights[i]):
            if (value): # 判斷不是空字串
                value = float(value)
                ws.cell(row=start_row + index, column=7, value=value)
            else : 
                ws.cell(row=start_row + index, column=7, value=value)
        # 把作法合併
        ws.merge_cells(start_row=start_row, start_column=8, end_row=last_row, end_column=8)
        # 最後把起始start_row 加上填的材料數
        start_row += ingredients_length

    # 在迴圈外 把餐別加上去 
    ws['A5'] = "午餐"
    ws.merge_cells(start_row=5, start_column=1, end_row=start_row-1, end_column=1)
   # endregion
    # -----------午餐菜單 ---------------------------

    # ----------- 午餐成本 (把食材篩選出來而已 一樣的就不重打)--------------------------
    # region 

    # 使用 itertools.chain 來展平列表 dish_ingredients 已經是 meal_1取出來的二維陣列[[],[],[]]
    dish_ingredients = list(chain(*dish_ingredients))

    # 取出不一樣的
    unique_ingredients = []
    for item in dish_ingredients:
        if item not in unique_ingredients:
            unique_ingredients.append(item)

    # 放入excel
    for index, value in enumerate(unique_ingredients):
            ws.cell(row=6 + index, column=10, value=value)
            
    # endregion
    # ----------- 午餐成本----------------------------------------------------------
    
    # ----------- 午餐營養成分 --------------------------
    # region 
    # 遍歷 S4~Z4  營養素分析
    for i in range(19, 27):

        # 如果all_nutrition_datas = None 如同server.py 初始化的 (早餐還沒按下計算營養按鈕)
        if all_nutrition_datas is None:
            print("午餐還沒按下計算營養按鈕 (key = '午餐' 沒發現)")
            break
        # 阿如果all_nutrition_datas["早點"] 有東西 那就執行下面
        else:
            # 去除 S4_value 中的所有空格和換行符：
            cleaned_value = ws.cell(row=4, column=i).value.replace("\n", "").replace(" ", "")
            # {早餐 : [all_nutrition_names , all_nutrition_values], 午餐 : [ [] , [] ]}
            try:
                # 找到S4"熱量 (kcal)" 所匹配到的 營養成分結果的index
                compare_name_index = all_nutrition_datas["午餐"][0].index(cleaned_value)
                # 把早餐的熱量 (or 其他的營養成分) 有多少抓過來
                this_nutrition_value = all_nutrition_datas["午餐"][1][compare_name_index]
                # 放入S5
                ws.cell(row=5, column=i,value = this_nutrition_value)
            except ValueError:
                ws.cell(row=5, column=i,value = "not found")
            except KeyError:
                print("KEYERROR : 午餐~~")
                break
    # 遍歷 S12~AA12 礦物質分析
    for i in range(19, 28):

        # 如果all_nutrition_datas = None 如同server.py 初始化的 (早餐還沒按下計算營養按鈕)
        if all_nutrition_datas is None:
            print("午餐還沒按下計算營養按鈕 (key = '午餐' 沒發現)")
            break
        # 阿如果all_nutrition_datas["早餐"] 有東西 那就執行下面
        else:
            # 去除 S4_value 中的所有空格和換行符：
            cleaned_value = ws.cell(row=12, column=i).value.replace("\n", "").replace(" ", "")
            # {早餐 : [all_nutrition_names , all_nutrition_values], 午餐 : [ [] , [] ]}
            try:
                # 找到S4"熱量 (kcal)" 所匹配到的 營養成分結果的index
                compare_name_index = all_nutrition_datas["午餐"][0].index(cleaned_value)
                # 把早餐的熱量 (or 其他的營養成分) 有多少抓過來
                this_nutrition_value = all_nutrition_datas["午餐"][1][compare_name_index]
                # 放入S5
                ws.cell(row=13, column=i,value = this_nutrition_value)
            except ValueError:
                ws.cell(row=13, column=i,value = "not found")
            except KeyError:
                print("KEYERROR : 午餐~~")
                break
    # 遍歷 S20~AA20 維生素分析
    for i in range(19, 28):

        # 如果all_nutrition_datas = None 如同server.py 初始化的 (早餐還沒按下計算營養按鈕)
        if all_nutrition_datas is None:
            print("午餐還沒按下計算營養按鈕 (key = '午餐' 沒發現)")
            break
        # 阿如果all_nutrition_datas["早餐"] 有東西 那就執行下面
        else:
            # 去除 S4_value 中的所有空格和換行符：
            cleaned_value = ws.cell(row=20, column=i).value.replace("\n", "").replace(" ", "")
            # {早餐 : [all_nutrition_names , all_nutrition_values], 午餐 : [ [] , [] ]}
            try:
                # 找到S4"熱量 (kcal)" 所匹配到的 營養成分結果的index
                compare_name_index = all_nutrition_datas["午餐"][0].index(cleaned_value)
                # 把早餐的熱量 (or 其他的營養成分) 有多少抓過來
                this_nutrition_value = all_nutrition_datas["午餐"][1][compare_name_index]
                # 放入S5
                ws.cell(row=21, column=i,value = this_nutrition_value)
            except ValueError:
                ws.cell(row=21, column=i,value = "not found")
            except KeyError:
                print("KEYERROR : 午餐~~")
                break

    # endregion
    # ----------- 午餐營養成分----------------------------------------------------------
    # endregion

    # region 午點
    # -----------午點菜單 ---------------------------
    # region 
    ws = wb["午點  菜單+成本"]
    meal_1 = dish_data['meal_4']

    # 使用列表解析找到找餐中所有以 "dish_" 開頭的鍵
    dish_keys = [key for key in meal_1.keys() if key.startswith("dish_")]

    # 計算找到的菜名的數量
    number_of_dishes = len(dish_keys)
    
    # 使用列表解析取得每一個dish_n的name class 等等屬性 放到陣列中
    dish_names = [dish["name"] for dish in meal_1.values()] #["name1",name2,name3...]
    dish_ingredients = [dish["ingredient"] for dish in meal_1.values()] # [[],[],[]...]
    dish_class_1s = [dish["class_1"] for dish in meal_1.values()]
    dish_class_2s = [dish["class_2"] for dish in meal_1.values()]
    dish_exs = [dish["ex"] for dish in meal_1.values()]
    dish_edible_weights = [dish["edible_weight"] for dish in meal_1.values()]
    # 8/27 如果沒有任何一道菜 (反正部會進到for迴圈)  ws.merge_cells(start_row=5, start_column=1, end_row=start_row-1, end_column=1) 使end_row > start_row=5
    if(len(dish_names) == 0):
        start_row =6 #為了等等被減1
    else:
        # 從B5開始一直往下田dish
        start_row = 5

    for i in range(number_of_dishes):
        # 開始一個一個放 (仙取的這個dish有多少材料)
        ingredients_length = len(dish_ingredients[i])
        # 從B5~ 材料數+5-1
        last_row = start_row + ingredients_length -1
        # 合併菜名 並 放入菜名
        ws.cell(row=start_row, column=2, value=dish_names[i])
        ws.merge_cells(start_row=start_row, start_column=2, end_row=last_row, end_column=2)
        # 放材料名稱 C = 3
        for index, value in enumerate(dish_ingredients[i]):
            ws.cell(row=start_row + index, column=3, value=value)
        # 放入類別1 D = 4
        for index, value in enumerate(dish_class_1s[i]):
            ws.cell(row=start_row + index, column=4, value=value)
        # 放類別2 E = 5
        for index, value in enumerate(dish_class_2s[i]):
            ws.cell(row=start_row + index, column=5, value=value)
        # 放EX F = 6
        for index, value in enumerate(dish_exs[i]):
            ws.cell(row=start_row + index, column=6, value=value)
        # 放可食重量 G = 7
        for index, value in enumerate(dish_edible_weights[i]):
            if (value): # 判斷不是空字串
                value = float(value)
                ws.cell(row=start_row + index, column=7, value=value)
            else : 
                ws.cell(row=start_row + index, column=7, value=value)
        # 把作法合併
        ws.merge_cells(start_row=start_row, start_column=8, end_row=last_row, end_column=8)
        # 最後把起始start_row 加上填的材料數
        start_row += ingredients_length

    # 在迴圈外 把餐別加上去 
    ws['A5'] = "午點"
    ws.merge_cells(start_row=5, start_column=1, end_row=start_row-1, end_column=1)
   # endregion
    # -----------午點菜單 ---------------------------

    # ----------- 午點成本 (把食材篩選出來而已 一樣的就不重打)--------------------------
    # region 

    # 使用 itertools.chain 來展平列表 dish_ingredients 已經是 meal_1取出來的二維陣列[[],[],[]]
    dish_ingredients = list(chain(*dish_ingredients))

    # 取出不一樣的
    unique_ingredients = []
    for item in dish_ingredients:
        if item not in unique_ingredients:
            unique_ingredients.append(item)

    # 放入excel
    for index, value in enumerate(unique_ingredients):
            ws.cell(row=6 + index, column=10, value=value)
            
    # endregion
    # ----------- 午點成本----------------------------------------------------------

     # ----------- 午點營養成分 --------------------------
    # region 
    # 遍歷 S4~Z4  營養素分析
    for i in range(19, 27):

        # 如果all_nutrition_datas = None 如同server.py 初始化的 (早餐還沒按下計算營養按鈕)
        if all_nutrition_datas is None:
            print("午點還沒按下計算營養按鈕 (key = '午點' 沒發現)")
            break
        # 阿如果all_nutrition_datas["早點"] 有東西 那就執行下面
        else:
            # 去除 S4_value 中的所有空格和換行符：
            cleaned_value = ws.cell(row=4, column=i).value.replace("\n", "").replace(" ", "")
            # {早餐 : [all_nutrition_names , all_nutrition_values], 午餐 : [ [] , [] ]}
            try:
                # 找到S4"熱量 (kcal)" 所匹配到的 營養成分結果的index
                compare_name_index = all_nutrition_datas["午點"][0].index(cleaned_value)
                # 把早餐的熱量 (or 其他的營養成分) 有多少抓過來
                this_nutrition_value = all_nutrition_datas["午點"][1][compare_name_index]
                # 放入S5
                ws.cell(row=5, column=i,value = this_nutrition_value)
            except ValueError:
                ws.cell(row=5, column=i,value = "not found")
            except KeyError:
                print("KEYERROR : 午點~~")
                break
    # 遍歷 S12~AA12 礦物質分析
    for i in range(19, 28):

        # 如果all_nutrition_datas = None 如同server.py 初始化的 (早餐還沒按下計算營養按鈕)
        if all_nutrition_datas is None:
            print("午點還沒按下計算營養按鈕 (key = '午點' 沒發現)")
            break
        # 阿如果all_nutrition_datas["早餐"] 有東西 那就執行下面
        else:
            # 去除 S4_value 中的所有空格和換行符：
            cleaned_value = ws.cell(row=12, column=i).value.replace("\n", "").replace(" ", "")
            # {早餐 : [all_nutrition_names , all_nutrition_values], 午餐 : [ [] , [] ]}
            try:
                # 找到S4"熱量 (kcal)" 所匹配到的 營養成分結果的index
                compare_name_index = all_nutrition_datas["午點"][0].index(cleaned_value)
                # 把早餐的熱量 (or 其他的營養成分) 有多少抓過來
                this_nutrition_value = all_nutrition_datas["午點"][1][compare_name_index]
                # 放入S5
                ws.cell(row=13, column=i,value = this_nutrition_value)
            except ValueError:
                ws.cell(row=13, column=i,value = "not found")
            except KeyError:
                print("KEYERROR : 午點~~")
                break
    # 遍歷 S20~AA20 維生素分析
    for i in range(19, 28):

        # 如果all_nutrition_datas = None 如同server.py 初始化的 (早餐還沒按下計算營養按鈕)
        if all_nutrition_datas is None:
            print("午點還沒按下計算營養按鈕 (key = '午點' 沒發現)")
            break
        # 阿如果all_nutrition_datas["早餐"] 有東西 那就執行下面
        else:
            # 去除 S4_value 中的所有空格和換行符：
            cleaned_value = ws.cell(row=20, column=i).value.replace("\n", "").replace(" ", "")
            # {早餐 : [all_nutrition_names , all_nutrition_values], 午餐 : [ [] , [] ]}
            try:
                # 找到S4"熱量 (kcal)" 所匹配到的 營養成分結果的index
                compare_name_index = all_nutrition_datas["午點"][0].index(cleaned_value)
                # 把早餐的熱量 (or 其他的營養成分) 有多少抓過來
                this_nutrition_value = all_nutrition_datas["午點"][1][compare_name_index]
                # 放入S5
                ws.cell(row=21, column=i,value = this_nutrition_value)
            except ValueError:
                ws.cell(row=21, column=i,value = "not found")
            except KeyError:
                print("KEYERROR : 午點~~")
                break

    # endregion
    # ----------- 午點營養成分----------------------------------------------------------
    # endregion
    
    # region 晚餐
    # -----------晚餐菜單 ---------------------------
    # region 
    ws = wb["晚餐  菜單+成本"]
    meal_1 = dish_data['meal_5']

    # 使用列表解析找到找餐中所有以 "dish_" 開頭的鍵
    dish_keys = [key for key in meal_1.keys() if key.startswith("dish_")]

    # 計算找到的菜名的數量
    number_of_dishes = len(dish_keys)

    # 使用列表解析取得每一個dish_n的name class 等等屬性 放到陣列中
    dish_names = [dish["name"] for dish in meal_1.values()] #["name1",name2,name3...]
    dish_ingredients = [dish["ingredient"] for dish in meal_1.values()] # [[],[],[]...]
    dish_class_1s = [dish["class_1"] for dish in meal_1.values()]
    dish_class_2s = [dish["class_2"] for dish in meal_1.values()]
    dish_exs = [dish["ex"] for dish in meal_1.values()]
    dish_edible_weights = [dish["edible_weight"] for dish in meal_1.values()]
    # 8/27 如果沒有任何一道菜 (反正部會進到for迴圈)  ws.merge_cells(start_row=5, start_column=1, end_row=start_row-1, end_column=1) 使end_row > start_row=5
    if(len(dish_names) == 0):
        start_row =6 #為了等等被減1
    else:
        # 從B5開始一直往下田dish
        start_row = 5

    for i in range(number_of_dishes):
        # 開始一個一個放 (仙取的這個dish有多少材料)
        ingredients_length = len(dish_ingredients[i])
        # 從B5~ 材料數+5-1
        last_row = start_row + ingredients_length -1
        # 合併菜名 並 放入菜名
        ws.cell(row=start_row, column=2, value=dish_names[i])
        ws.merge_cells(start_row=start_row, start_column=2, end_row=last_row, end_column=2)
        # 放材料名稱 C = 3
        for index, value in enumerate(dish_ingredients[i]):
            ws.cell(row=start_row + index, column=3, value=value)
        # 放入類別1 D = 4
        for index, value in enumerate(dish_class_1s[i]):
            ws.cell(row=start_row + index, column=4, value=value)
        # 放類別2 E = 5
        for index, value in enumerate(dish_class_2s[i]):
            ws.cell(row=start_row + index, column=5, value=value)
        # 放EX F = 6
        for index, value in enumerate(dish_exs[i]):
            ws.cell(row=start_row + index, column=6, value=value)
        # 放可食重量 G = 7
        for index, value in enumerate(dish_edible_weights[i]):
            if (value): # 判斷不是空字串
                value = float(value)
                ws.cell(row=start_row + index, column=7, value=value)
            else : 
                ws.cell(row=start_row + index, column=7, value=value)
        # 把作法合併
        ws.merge_cells(start_row=start_row, start_column=8, end_row=last_row, end_column=8)
        # 最後把起始start_row 加上填的材料數
        start_row += ingredients_length

    # 在迴圈外 把餐別加上去 
    ws['A5'] = "晚餐"
    ws.merge_cells(start_row=5, start_column=1, end_row=start_row-1, end_column=1)
   # endregion
    # -----------晚餐菜單 ---------------------------

    # ----------- 晚餐成本 (把食材篩選出來而已 一樣的就不重打)--------------------------
    # region 

    # 使用 itertools.chain 來展平列表 dish_ingredients 已經是 meal_1取出來的二維陣列[[],[],[]]
    dish_ingredients = list(chain(*dish_ingredients))

    # 取出不一樣的
    unique_ingredients = []
    for item in dish_ingredients:
        if item not in unique_ingredients:
            unique_ingredients.append(item)

    # 放入excel
    for index, value in enumerate(unique_ingredients):
            ws.cell(row=6 + index, column=10, value=value)
            
    # endregion
    # ----------- 晚餐成本----------------------------------------------------------
        
    # ----------- 晚餐營養成分 --------------------------
    # region 
    # 遍歷 S4~Z4  營養素分析
    for i in range(19, 27):

        # 如果all_nutrition_datas = None 如同server.py 初始化的 (早餐還沒按下計算營養按鈕)
        if all_nutrition_datas is None:
            print("晚餐還沒按下計算營養按鈕 (key = '晚餐' 沒發現)")
            break
        # 阿如果all_nutrition_datas["早點"] 有東西 那就執行下面
        else:
            # 去除 S4_value 中的所有空格和換行符：
            cleaned_value = ws.cell(row=4, column=i).value.replace("\n", "").replace(" ", "")
            # {早餐 : [all_nutrition_names , all_nutrition_values], 午餐 : [ [] , [] ]}
            try:
                # 找到S4"熱量 (kcal)" 所匹配到的 營養成分結果的index
                compare_name_index = all_nutrition_datas["晚餐"][0].index(cleaned_value)
                # 把早餐的熱量 (or 其他的營養成分) 有多少抓過來
                this_nutrition_value = all_nutrition_datas["晚餐"][1][compare_name_index]
                # 放入S5
                ws.cell(row=5, column=i,value = this_nutrition_value)
            except ValueError:
                ws.cell(row=5, column=i,value = "not found")
            except KeyError:
                print("KEYERROR : 晚餐~~")
                break
    # 遍歷 S12~AA12 礦物質分析
    for i in range(19, 28):

        # 如果all_nutrition_datas = None 如同server.py 初始化的 (早餐還沒按下計算營養按鈕)
        if all_nutrition_datas is None:
            print("晚餐還沒按下計算營養按鈕 (key = '晚餐' 沒發現)")
            break
        # 阿如果all_nutrition_datas["早餐"] 有東西 那就執行下面
        else:
            # 去除 S4_value 中的所有空格和換行符：
            cleaned_value = ws.cell(row=12, column=i).value.replace("\n", "").replace(" ", "")
            # {早餐 : [all_nutrition_names , all_nutrition_values], 午餐 : [ [] , [] ]}
            try:
                # 找到S4"熱量 (kcal)" 所匹配到的 營養成分結果的index
                compare_name_index = all_nutrition_datas["晚餐"][0].index(cleaned_value)
                # 把早餐的熱量 (or 其他的營養成分) 有多少抓過來
                this_nutrition_value = all_nutrition_datas["晚餐"][1][compare_name_index]
                # 放入S5
                ws.cell(row=13, column=i,value = this_nutrition_value)
            except ValueError:
                ws.cell(row=13, column=i,value = "not found")
            except KeyError:
                print("KEYERROR : 晚餐~~")
                break
    # 遍歷 S20~AA20 維生素分析
    for i in range(19, 28):

        # 如果all_nutrition_datas = None 如同server.py 初始化的 (早餐還沒按下計算營養按鈕)
        if all_nutrition_datas is None:
            print("晚餐還沒按下計算營養按鈕 (key = '晚餐' 沒發現)")
            break
        # 阿如果all_nutrition_datas["早餐"] 有東西 那就執行下面
        else:
            # 去除 S4_value 中的所有空格和換行符：
            cleaned_value = ws.cell(row=20, column=i).value.replace("\n", "").replace(" ", "")
            # {早餐 : [all_nutrition_names , all_nutrition_values], 午餐 : [ [] , [] ]}
            try:
                # 找到S4"熱量 (kcal)" 所匹配到的 營養成分結果的index
                compare_name_index = all_nutrition_datas["晚餐"][0].index(cleaned_value)
                # 把早餐的熱量 (or 其他的營養成分) 有多少抓過來
                this_nutrition_value = all_nutrition_datas["晚餐"][1][compare_name_index]
                # 放入S5
                ws.cell(row=21, column=i,value = this_nutrition_value)
            except ValueError:
                ws.cell(row=21, column=i,value = "not found")
            except KeyError:
                print("KEYERROR : 晚餐~~")
                break

    # endregion
    # ----------- 晚餐營養成分----------------------------------------------------------
    # endregion

    # region 晚點
    # -----------晚點菜單 ---------------------------
    # region 
    ws = wb["晚點  菜單+成本"]
    meal_1 = dish_data['meal_6']

    # 使用列表解析找到找餐中所有以 "dish_" 開頭的鍵
    dish_keys = [key for key in meal_1.keys() if key.startswith("dish_")]

    # 計算找到的菜名的數量
    number_of_dishes = len(dish_keys)

    # 使用列表解析取得每一個dish_n的name class 等等屬性 放到陣列中
    dish_names = [dish["name"] for dish in meal_1.values()] #["name1",name2,name3...]
    dish_ingredients = [dish["ingredient"] for dish in meal_1.values()] # [[],[],[]...]
    dish_class_1s = [dish["class_1"] for dish in meal_1.values()]
    dish_class_2s = [dish["class_2"] for dish in meal_1.values()]
    dish_exs = [dish["ex"] for dish in meal_1.values()]
    dish_edible_weights = [dish["edible_weight"] for dish in meal_1.values()]
    # 8/27 如果沒有任何一道菜 (反正部會進到for迴圈)  ws.merge_cells(start_row=5, start_column=1, end_row=start_row-1, end_column=1) 使end_row > start_row=5
    if(len(dish_names) == 0):
        start_row =6 #為了等等被減1
    else:
        # 從B5開始一直往下田dish
        start_row = 5

    for i in range(number_of_dishes):
        # 開始一個一個放 (仙取的這個dish有多少材料)
        ingredients_length = len(dish_ingredients[i])
        # 從B5~ 材料數+5-1
        last_row = start_row + ingredients_length -1
        # 合併菜名 並 放入菜名
        ws.cell(row=start_row, column=2, value=dish_names[i])
        ws.merge_cells(start_row=start_row, start_column=2, end_row=last_row, end_column=2)
        # 放材料名稱 C = 3
        for index, value in enumerate(dish_ingredients[i]):
            ws.cell(row=start_row + index, column=3, value=value)
        # 放入類別1 D = 4
        for index, value in enumerate(dish_class_1s[i]):
            ws.cell(row=start_row + index, column=4, value=value)
        # 放類別2 E = 5
        for index, value in enumerate(dish_class_2s[i]):
            ws.cell(row=start_row + index, column=5, value=value)
        # 放EX F = 6
        for index, value in enumerate(dish_exs[i]):
            ws.cell(row=start_row + index, column=6, value=value)
        # 放可食重量 G = 7
        for index, value in enumerate(dish_edible_weights[i]):
            if (value): # 判斷不是空字串
                value = float(value)
                ws.cell(row=start_row + index, column=7, value=value)
            else : 
                ws.cell(row=start_row + index, column=7, value=value)
        # 把作法合併
        ws.merge_cells(start_row=start_row, start_column=8, end_row=last_row, end_column=8)
        # 最後把起始start_row 加上填的材料數
        start_row += ingredients_length

    # 在迴圈外 把餐別加上去 
    ws['A5'] = "晚點"
    ws.merge_cells(start_row=5, start_column=1, end_row=start_row-1, end_column=1)
   # endregion
    # -----------晚點菜單 ---------------------------

    # ----------- 晚點成本 (把食材篩選出來而已 一樣的就不重打)--------------------------
    # region 

    # 使用 itertools.chain 來展平列表 dish_ingredients 已經是 meal_1取出來的二維陣列[[],[],[]]
    dish_ingredients = list(chain(*dish_ingredients))

    # 取出不一樣的
    unique_ingredients = []
    for item in dish_ingredients:
        if item not in unique_ingredients:
            unique_ingredients.append(item)

    # 放入excel
    for index, value in enumerate(unique_ingredients):
            ws.cell(row=6 + index, column=10, value=value)
            
    # endregion
    # ----------- 晚點成本----------------------------------------------------------
    
    # ----------- 晚點營養成分 --------------------------
    # region 
    # 遍歷 S4~Z4  營養素分析
    for i in range(19, 27):

        # 如果all_nutrition_datas = None 如同server.py 初始化的 (早餐還沒按下計算營養按鈕)
        if all_nutrition_datas is None:
            print("晚點還沒按下計算營養按鈕 (key = '晚點' 沒發現)")
            break
        # 阿如果all_nutrition_datas["早點"] 有東西 那就執行下面
        else:
            # 去除 S4_value 中的所有空格和換行符：
            cleaned_value = ws.cell(row=4, column=i).value.replace("\n", "").replace(" ", "")
            # {早餐 : [all_nutrition_names , all_nutrition_values], 午餐 : [ [] , [] ]}
            try:
                # 找到S4"熱量 (kcal)" 所匹配到的 營養成分結果的index
                compare_name_index = all_nutrition_datas["晚點"][0].index(cleaned_value)
                # 把早餐的熱量 (or 其他的營養成分) 有多少抓過來
                this_nutrition_value = all_nutrition_datas["晚點"][1][compare_name_index]
                # 放入S5
                ws.cell(row=5, column=i,value = this_nutrition_value)
            except ValueError:
                ws.cell(row=5, column=i,value = "not found")
            except KeyError:
                print("KEYERROR : 晚點~~")
                break
    # 遍歷 S12~AA12 礦物質分析
    for i in range(19, 28):

        # 如果all_nutrition_datas = None 如同server.py 初始化的 (早餐還沒按下計算營養按鈕)
        if all_nutrition_datas is None:
            print("晚點還沒按下計算營養按鈕 (key = '晚點' 沒發現)")
            break
        # 阿如果all_nutrition_datas["早餐"] 有東西 那就執行下面
        else:
            # 去除 S4_value 中的所有空格和換行符：
            cleaned_value = ws.cell(row=12, column=i).value.replace("\n", "").replace(" ", "")
            # {早餐 : [all_nutrition_names , all_nutrition_values], 午餐 : [ [] , [] ]}
            try:
                # 找到S4"熱量 (kcal)" 所匹配到的 營養成分結果的index
                compare_name_index = all_nutrition_datas["晚點"][0].index(cleaned_value)
                # 把早餐的熱量 (or 其他的營養成分) 有多少抓過來
                this_nutrition_value = all_nutrition_datas["晚點"][1][compare_name_index]
                # 放入S5
                ws.cell(row=13, column=i,value = this_nutrition_value)
            except ValueError:
                ws.cell(row=13, column=i,value = "not found")
            except KeyError:
                print("KEYERROR : 晚點~~")
                break
    # 遍歷 S20~AA20 維生素分析
    for i in range(19, 28):

        # 如果all_nutrition_datas = None 如同server.py 初始化的 (早餐還沒按下計算營養按鈕)
        if all_nutrition_datas is None:
            print("晚點還沒按下計算營養按鈕 (key = '晚點' 沒發現)")
            break
        # 阿如果all_nutrition_datas["早餐"] 有東西 那就執行下面
        else:
            # 去除 S4_value 中的所有空格和換行符：
            cleaned_value = ws.cell(row=20, column=i).value.replace("\n", "").replace(" ", "")
            # {早餐 : [all_nutrition_names , all_nutrition_values], 午餐 : [ [] , [] ]}
            try:
                # 找到S4"熱量 (kcal)" 所匹配到的 營養成分結果的index
                compare_name_index = all_nutrition_datas["晚點"][0].index(cleaned_value)
                # 把早餐的熱量 (or 其他的營養成分) 有多少抓過來
                this_nutrition_value = all_nutrition_datas["晚點"][1][compare_name_index]
                # 放入S5
                ws.cell(row=21, column=i,value = this_nutrition_value)
            except ValueError:
                ws.cell(row=21, column=i,value = "not found")
            except KeyError:
                print("KEYERROR : 晚點~~")
                break

    # endregion
    # ----------- 晚點營養成分----------------------------------------------------------
    # endregion

# endregion
    ## ------------- 菜單-------------------------------------



    wb.save('自動化表格test.xlsx')


    

   
