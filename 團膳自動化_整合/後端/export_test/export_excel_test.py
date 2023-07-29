from flask import  request
from openpyxl import Workbook, load_workbook ## 新增到excel所需要的模組 (21行~53行)
from openpyxl.styles import Alignment ## 置中用的
from openpyxl.utils import get_column_letter ## 把數字轉字母
from openpyxl.styles import PatternFill ##填充顏色
from openpyxl.formula.translate import Translator ##寫入 翻譯成excel的公式

# // 回傳的東東 舉例
#     // {
#     //     class_inputs: (3) ['乳品', '豆、魚、蛋、肉類', '蔬菜類']
#     //     name_inputs: (3) ['寶乃', '巴沙魚', '高麗菜']
#     //     needness_inputs: (3) ['23', '43', '56']
#     //     unit_inputs: (3) ['台斤', '公斤', '公升']
#     //     prices_db: (3) ['80', '80', '20']
#     //     units_db: (3) ['台斤', '台斤', '公克']
#     //     results: (3) ['1840.000', '5733.333', '1120000.000']
#     //     total_price: 1127573.333
#     // }
# //




##-----------把前端傳來的東東存在變數裡--------------
data_json = {'class_inputs': ['乳品', '豆、魚、蛋、肉類', '全榖雜糧類', '蔬菜類', '水果類', '蔬菜類', '豆、魚、蛋、肉類', '蔬菜類', '全榖雜糧類'], 'name_inputs': ['寶乃', '巴沙魚', '大米', '高麗菜', '蘋果', '小白菜', '紅肉', '小白菜', '小米'], 'needness_inputs': ['23', '24', '54', '65', '23', '34', '45', '56', '76'], 'unit_inputs': ['公克', '公克', '公克', '公克', '公克', '公克', '公克', '公克', '公克'], 'prices_db': ['50', '50', '50', '50', '50', '50', '50', '50', '67'], 'units_db': ['台斤', '台斤', '台斤', '台斤', '台斤', '台斤', '台斤', '台斤', '公克'], 'results': ['1.917', '2.000', '4.500', '5.417', '1.917', '2.833', '3.750', '4.667', '5092.000'], 'total_price': 5119.001}

class_inputs = data_json['class_inputs']
name_inputs = data_json['name_inputs']
needness_inputs = data_json['needness_inputs']
unit_inputs = data_json['unit_inputs']
prices_db = data_json['prices_db']
units_db = data_json['units_db']
results = data_json['results']
total_price = data_json['total_price']

print(data_json)
##-----------把前端傳來的東東存在變數裡--------------

# 建立一個新的Excel檔案
wb = Workbook()

# 建立名為"早餐"的工作表
ws = wb.create_sheet('早餐')

# 合併儲存格
ws.merge_cells('A1:G1')  # 合併A1到G1的儲存格

# 寫入數據到合併後的儲存格中，僅需寫入合併儲存格的左上角位置即可
ws['A1'] = '早餐'

# 把設定置中寫成變數~ wrapText=True 打開自動換行
align_center = Alignment(horizontal='center', vertical='center' , wrapText=True)

#早餐置中
ws['A1'].alignment = align_center

#填入標題
ws['A2'] = '食材名稱'
ws['B2'] = '可食重量'
ws['C2'] = '採買重量     (克)'
ws['D2'] = '預估單價     (元/公斤)'
ws['E2'] = '預估價格     (元)'
ws['F2'] = '實際單價     (元/公斤)'
ws['G2'] = '實際價格     (元/公斤)'

# 將第2行的高度設定為30（單位是為像素，30像素約等於1.38行高度的標準高度）
ws.row_dimensions[2].height = 30

# 设置连续列列宽：
for column_num in range(1,8):  # 注意，列序数从1开始，但必须转变为A\B等字母
    cl = get_column_letter(column_num)  # 把列序数转变为字母
    ws.column_dimensions[cl].width = 15 # 常见列宽8.47,约1.8cm。  如果15，约3.1cm

# 連續把第二行設置中
for column_num in range(1,8):  # 注意，列序数从1开始，但必须转变为A\B等字母
    cl = get_column_letter(column_num)  # 把列序数转变为字母
    ws[cl+'2'].alignment = align_center

# 設定A1儲存格(早餐)的填充顏色為黃色
yellow_fill = PatternFill(start_color='F7E091', end_color='F7E091', fill_type='solid') #start:end 漸層
ws['A1'].fill = yellow_fill

# 設定A2~G2儲存格的填充顏色為淡黃色
light_yellow_fill = PatternFill(start_color='F7E9B3', end_color='F7E9B3', fill_type='solid') #start:end 漸層
# copy设置连续列列宽 連續設置列顏色：
for column_num in range(1,8):  # 注意，列序数从1开始，但必须转变为A\B等字母
    cl = get_column_letter(column_num)  # 把列序数转变为字母
    ws[cl+'2'].fill = light_yellow_fill



# <<A 食材名稱>>  列使用for迴圈將名稱元素逐一放入A列
for index, name in enumerate(name_inputs, start=3): #A3開始
    cell = 'A' + str(index)  # 這裡使用列號和索引值組合來獲取儲存格的位置
    ws[cell] = name
    ws[cell].alignment = align_center ## 順便置中

# <<B 可食重量>>  列使用for迴圈將名稱元素逐一放入B列
for index, name in enumerate(name_inputs, start=3): #A3開始
    cell = 'B' + str(index)  # 這裡使用列號和索引值組合來獲取儲存格的位置
    ws[cell] = ''
    ws[cell].alignment = align_center ## 順便置中

# <<C 採買重量>>  列使用for迴圈將名稱元素逐一放入C列
for index, needness in enumerate(needness_inputs, start=3): #A3開始
    cell = 'C' + str(index)  # 這裡使用列號和索引值組合來獲取儲存格的位置
    ws[cell] = needness
    ws[cell].alignment = align_center ## 順便置中

# <<D 預估單價>>  列使用for迴圈將名稱元素逐一放入D列
for index, prices_db in enumerate(prices_db, start=3): #A3開始
    cell = 'D' + str(index)  # 這裡使用列號和索引值組合來獲取儲存格的位置
    ws[cell] = prices_db
    ws[cell].alignment = align_center ## 順便置中

# <<E> 預估價格>  列使用for迴圈將名稱元素逐一放入E列
for index, result in enumerate(results, start=3): #A3開始
    cell = 'E' + str(index)  # 這裡使用列號和索引值組合來獲取儲存格的位置
    ws[cell] = result
    ws[cell].alignment = align_center ## 順便置中

# <<F 實際單價>>  列使用for迴圈將名稱元素逐一放入F列
for index, name in enumerate(name_inputs, start=3): #A3開始
    cell = 'F' + str(index)  # 這裡使用列號和索引值組合來獲取儲存格的位置
    ws[cell] = ''
    ws[cell].alignment = align_center ## 順便置中

# <<G 實際價格>>  列使用for迴圈將名稱元素逐一放入G列# 將整排G的儲存格填入 C*F
for index, name in enumerate(name_inputs, start=3): #A3開始
    # 這裡使用列號和索引值組合來獲取儲存格的位置
    cell = 'G' + str(index)  
    # 相乘公式
    ws[cell] = '= C'+str(index) + '*F' + str(index) #"=C{index}*F{index}" 
    ## 順便置中
    ws[cell].alignment = align_center ## 順便置中




# 總價
# 使用len()函數獲取列表的長度
length = len(name_inputs)
total_price_cell = str(length+3) ##從A3開始
#"總價"在A的最下面
ws['A'+ total_price_cell ] = '總價'
ws['A'+ total_price_cell ].alignment = align_center ## 順便置中
# 預估價格的總價
ws['E' + total_price_cell] = total_price
ws['E' + total_price_cell].alignment = align_center ## 順便置中
#實際價格的總價
# 在G最下面的儲存格中放入動態的相加公式，假設G11是動態的，即其值由其他地方得出
# str(length+2)是最後的職
GP_cell = 'G'+ total_price_cell #GP g排cell
ws[GP_cell] = '=SUM(G4:G'+str(length+2)+')'  #'=  SUM(G4:G{str(length+2)}) '
ws[GP_cell].alignment = align_center ## 順便置中





wb.save('test.xlsx')


